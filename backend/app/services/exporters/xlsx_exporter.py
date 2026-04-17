"""XLSX exporter with styled headers and per-data-type tabs."""

from __future__ import annotations

import os
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Alignment, Side
from openpyxl.utils import get_column_letter


# Financial theme colors — dark blue headers
HEADER_BG_COLOR = "1B3A5C"
HEADER_FONT_COLOR = "FFFFFF"

# Trial Balance styling
SECTION_HEADER_BG = "2C5282"
SECTION_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_HEADER_FILL = PatternFill(start_color=SECTION_HEADER_BG, end_color=SECTION_HEADER_BG, fill_type="solid")
SUBTOTAL_BG = "E2E8F0"
SUBTOTAL_FILL = PatternFill(start_color=SUBTOTAL_BG, end_color=SUBTOTAL_BG, fill_type="solid")
SUBTOTAL_FONT = Font(bold=True, size=11)
GRAND_TOTAL_BG = "1B3A5C"
GRAND_TOTAL_FILL = PatternFill(start_color=GRAND_TOTAL_BG, end_color=GRAND_TOTAL_BG, fill_type="solid")
GRAND_TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER_BOTTOM = Border(bottom=Side(style="thin"))

PASS_FILL = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
PASS_FONT = Font(color="1B7D3F", size=11)
FAIL_FONT = Font(color="C5221F", bold=True, size=11)


def export_xlsx(
    data_types: dict[str, list[dict[str, Any]]],
    output_dir: str,
    base_name: str = "export",
    validation_checks: list[dict[str, Any]] | None = None,
    trial_balance: tuple[list[dict[str, Any]], list[dict[str, Any]]] | None = None,
) -> str:
    """Export data types to a single .xlsx file with one tab per data type.

    Returns the filename of the output file.
    """
    filename = f"{base_name}.xlsx"
    filepath = os.path.join(output_dir, filename)

    wb = Workbook()
    # Remove the default sheet
    wb.remove(wb.active)

    header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
    header_fill = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for dt_name, rows in data_types.items():
        # Sheet name max 31 chars in Excel
        sheet_name = dt_name[:31]
        ws = wb.create_sheet(title=sheet_name)

        if not rows:
            continue

        # Collect all columns across all rows
        columns = _all_columns(rows)

        # Write header row
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows with explicit types
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row.get(col_name, "")
                if value is None:
                    value = ""
                # Try to convert numeric strings to numbers for proper Excel formatting
                cell_value = _typed_value(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Auto-width columns based on content
        for col_idx, col_name in enumerate(columns, 1):
            max_width = len(str(col_name))
            # Sample first 100 rows to determine width
            for row_idx in range(2, min(len(rows) + 2, 102)):
                cell_val = ws.cell(row=row_idx, column=col_idx).value
                if cell_val is not None:
                    max_width = max(max_width, len(str(cell_val)))
            # Cap at 50 chars, add padding
            adjusted_width = min(max_width + 3, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Add Validation sheet if checks are provided
    if validation_checks:
        vs = wb.create_sheet(title="Validation", index=0)
        val_headers = ["Section", "Check", "Declared (XAF)", "Computed", "Result"]
        val_header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
        val_header_fill = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type="solid")

        for col_idx, h in enumerate(val_headers, 1):
            cell = vs.cell(row=1, column=col_idx, value=h)
            cell.font = val_header_font
            cell.fill = val_header_fill
            cell.alignment = header_alignment

        for row_idx, vc in enumerate(validation_checks, 2):
            passed = vc.get("passed", False)
            row_fill = PASS_FILL if passed else FAIL_FILL
            row_font = PASS_FONT if passed else FAIL_FONT
            result_text = "PASS" if passed else "FAIL"

            values = [
                vc.get("section", ""),
                vc.get("check", ""),
                vc.get("declared", ""),
                vc.get("computed", ""),
                result_text,
            ]
            for col_idx, val in enumerate(values, 1):
                cell = vs.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = row_fill
                if col_idx == 5:
                    cell.font = row_font

        for col_idx in range(1, 6):
            vs.column_dimensions[get_column_letter(col_idx)].width = 22

    # Add Trial Balance sheet if data is provided
    if trial_balance is not None:
        balance_rows, pl_rows = trial_balance
        if balance_rows or pl_rows:
            _write_trial_balance_sheet(wb, balance_rows, pl_rows, header_font, header_fill, header_alignment)

    # Ensure at least one sheet exists
    if len(wb.sheetnames) == 0:
        wb.create_sheet(title="Empty")

    wb.save(filepath)
    return filename


def _all_columns(rows: list[dict[str, Any]]) -> list[str]:
    """Collect all unique column names across all rows, preserving order."""
    seen: dict[str, None] = {}
    for row in rows:
        for key in row:
            if key not in seen:
                seen[key] = None
    return list(seen.keys())


def _typed_value(value: Any) -> Any:
    """Convert value to appropriate Excel type.

    Strings are kept as strings (explicit type).
    Numeric strings that look like amounts are converted to float.
    """
    if not isinstance(value, str):
        return value
    if value == "":
        return ""
    # Try float conversion for numeric-looking values
    try:
        fval = float(value)
        # Preserve integer appearance
        if fval == int(fval) and "." not in value:
            return int(fval)
        return fval
    except (ValueError, OverflowError):
        return str(value)


# ---------------------------------------------------------------------------
# Trial Balance sheet
# ---------------------------------------------------------------------------

# Column layout for the trial balance
_TB_HEADERS = [
    "Account", "Description",
    "Opening Debit", "Opening Credit",
    "Debit", "Credit",
    "Year Debit", "Year Credit",
    "End Debit", "End Credit",
]

_TB_NUM_FMT = '#,##0.00'


def _write_trial_balance_sheet(
    wb: Workbook,
    balance_rows: list[dict[str, Any]],
    pl_rows: list[dict[str, Any]],
    header_font: Font,
    header_fill: PatternFill,
    header_alignment: Alignment,
) -> None:
    """Write the Trial Balance tab at position 1 (after Validation).

    Uses Excel formulas:
    - Data rows: Year = Opening + Mutations, End = net of Year (IF formula)
    - Subtotals: SUM over data rows
    - Grand total: sum of subtotal rows
    """
    ws = wb.create_sheet(title="Trial Balance", index=1)

    # Column widths
    col_widths = [14, 36, 16, 16, 16, 16, 16, 16, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Write main header row
    for col_idx, h in enumerate(_TB_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    current_row = 2
    subtotal_rows: list[int] = []  # track subtotal row numbers for grand total

    # Balance section
    if balance_rows:
        current_row, sub_row = _write_tb_section(ws, current_row, "Balance Sheet Accounts", balance_rows)
        subtotal_rows.append(sub_row)

    # P&L section
    if pl_rows:
        current_row, sub_row = _write_tb_section(ws, current_row, "Profit & Loss Accounts", pl_rows)
        subtotal_rows.append(sub_row)

    # Grand total — sums the subtotal rows
    _write_tb_grand_total(ws, current_row, subtotal_rows, GRAND_TOTAL_FILL, GRAND_TOTAL_FONT)

    # Freeze top row
    ws.freeze_panes = "A2"


def _write_tb_section(
    ws,
    start_row: int,
    section_title: str,
    rows: list[dict[str, Any]],
) -> tuple[int, int]:
    """Write a section (Balance or P&L).

    Returns (next_row, subtotal_row) so the grand total can reference subtotals.
    """
    row_num = start_row

    # Section header row
    cell = ws.cell(row=row_num, column=1, value=section_title)
    cell.font = SECTION_HEADER_FONT
    cell.fill = SECTION_HEADER_FILL
    for col in range(2, 11):
        ws.cell(row=row_num, column=col).fill = SECTION_HEADER_FILL
    row_num += 1

    # Value fields written as literal numbers (columns C–F)
    value_fields = ["ob_debit", "ob_credit", "mut_debit", "mut_credit"]
    data_first_row = row_num

    for r in rows:
        ws.cell(row=row_num, column=1, value=r["accID"])
        ws.cell(row=row_num, column=2, value=r["accDesc"])

        # Columns C(3)–F(6): Opening Debit, Opening Credit, Debit, Credit
        for col_idx, field_name in enumerate(value_fields, 3):
            val = r.get(field_name, 0.0)
            cell = ws.cell(row=row_num, column=col_idx)
            if val:
                cell.value = val
            cell.number_format = _TB_NUM_FMT

        # Column G(7): Year Debit = Opening Debit + Mutations Debit
        cell = ws.cell(row=row_num, column=7)
        cell.value = f"=C{row_num}+E{row_num}"
        cell.number_format = _TB_NUM_FMT

        # Column H(8): Year Credit = Opening Credit + Mutations Credit
        cell = ws.cell(row=row_num, column=8)
        cell.value = f"=D{row_num}+F{row_num}"
        cell.number_format = _TB_NUM_FMT

        # Column I(9): End Debit = IF(Year Debit > Year Credit, difference, 0)
        cell = ws.cell(row=row_num, column=9)
        cell.value = f"=IF(G{row_num}-H{row_num}>0,G{row_num}-H{row_num},0)"
        cell.number_format = _TB_NUM_FMT

        # Column J(10): End Credit = IF(Year Credit > Year Debit, difference, 0)
        cell = ws.cell(row=row_num, column=10)
        cell.value = f"=IF(H{row_num}-G{row_num}>0,H{row_num}-G{row_num},0)"
        cell.number_format = _TB_NUM_FMT

        row_num += 1

    data_last_row = row_num - 1

    # Subtotal row with SUM formulas over data rows
    subtotal_row = row_num
    _write_tb_subtotal(ws, subtotal_row, f"Subtotal {section_title}",
                       data_first_row, data_last_row, SUBTOTAL_FILL, SUBTOTAL_FONT)
    row_num += 1

    # Blank separator row
    row_num += 1

    return row_num, subtotal_row


def _write_tb_subtotal(
    ws,
    row_num: int,
    label: str,
    data_first: int,
    data_last: int,
    fill: PatternFill,
    font: Font,
) -> None:
    """Write a subtotal row with SUM formulas over a data range."""
    ws.cell(row=row_num, column=1, value=label).font = font
    ws.cell(row=row_num, column=1).fill = fill
    ws.cell(row=row_num, column=2).fill = fill

    # SUM formulas for columns C(3) through J(10)
    for col_idx in range(3, 11):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = f"=SUM({col_letter}{data_first}:{col_letter}{data_last})"
        cell.number_format = _TB_NUM_FMT
        cell.font = font
        cell.fill = fill


def _write_tb_grand_total(
    ws,
    row_num: int,
    subtotal_rows: list[int],
    fill: PatternFill,
    font: Font,
) -> None:
    """Write a grand total row that sums the subtotal rows."""
    ws.cell(row=row_num, column=1, value="Grand Total").font = font
    ws.cell(row=row_num, column=1).fill = fill
    ws.cell(row=row_num, column=2).fill = fill

    for col_idx in range(3, 11):
        col_letter = get_column_letter(col_idx)
        # Build formula like =C5+C12 referencing the subtotal rows
        parts = [f"{col_letter}{r}" for r in subtotal_rows]
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = f"={'+'.join(parts)}"
        cell.number_format = _TB_NUM_FMT
        cell.font = font
        cell.fill = fill
